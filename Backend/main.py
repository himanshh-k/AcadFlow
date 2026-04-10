import io
import math
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict
from ortools.sat.python import cp_model

# Import openpyxl components for advanced Excel formatting
import openpyxl.utils
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

app = FastAPI(title="AcadFlow: Intelligent Academic Scheduling and Resource Management")

# ==========================================
# 1. DATABASE MODELS (API Payloads)
# ==========================================

class Course(BaseModel):
    id: str
    name: str
    teachers: List[str]
    section: str
    hours: int
    is_lab: bool = False
    elective_group: Optional[str] = None

class TimetableRequest(BaseModel):
    num_days: int = 6
    num_periods: int = 8  
    sections: List[str]
    teachers: List[str]
    rooms: List[str]
    courses: List[Course]

class ScheduledClass(BaseModel):
    day: int
    period: int
    room: str
    course_name: str
    teachers: List[str]
    is_recess: bool = False

class TimetableResponse(BaseModel):
    status: str
    schedule: Optional[Dict[str, List[ScheduledClass]]] = None
    message: str

class RescheduleRequest(BaseModel):
    current_schedule: Dict[str, List[ScheduledClass]]
    target_day: int 
    course: Course
    all_rooms: List[str]
    num_periods: int = 8

class ExportRequest(BaseModel):
    schedule: Dict[str, List[ScheduledClass]]
    num_days: int = 6
    num_periods: int = 8

# NEW: Extra Class Request Model
class ExtraClassRequest(BaseModel):
    current_schedule: Dict[str, List[ScheduledClass]]
    course_id: str
    section: str
    all_courses: List[Course]  # Needed to fetch teacher and lab info
    all_rooms: List[str]
    target_day: Optional[int] = None # Optional: If provided, only searches this day. If null, searches the whole week.
    num_days: int = 6
    num_periods: int = 8

# ==========================================
# 2. THE AI SOLVER CORE (OR-Tools)
# ==========================================

def generate_schedule(data: TimetableRequest) -> dict:
    model = cp_model.CpModel()
    schedule = {}
    recess = {}
    sec_active_day = {}
    
    # 1. Initialize Variables using Internal Index
    for c_idx, c in enumerate(data.courses):
        for r_idx in range(len(data.rooms)):
            for d in range(data.num_days):
                for p in range(data.num_periods):
                    schedule[(c_idx, r_idx, d, p)] = model.NewBoolVar(f"c{c_idx}_r{r_idx}_d{d}_p{p}")

    for section in data.sections:
        for d in range(data.num_days):
            for p in range(data.num_periods):
                recess[(section, d, p)] = model.NewBoolVar(f"rec_{section}_d{d}_p{p}")

    # 2. Define "Day Off" logic per section
    for section in data.sections:
        c_indices = [i for i, c in enumerate(data.courses) if c.section == section]
        for d in range(data.num_days):
            is_active = model.NewBoolVar(f'sec_{section}_d{d}_active')
            total_classes_day = sum(
                schedule[(i, r_idx, d, p)] 
                for i in c_indices for r_idx in range(len(data.rooms)) for p in range(data.num_periods)
            )
            model.Add(total_classes_day > 0).OnlyEnforceIf(is_active)
            model.Add(total_classes_day == 0).OnlyEnforceIf(is_active.Not())
            sec_active_day[(section, d)] = is_active

    # --- ADVANCED CONSTRAINTS ---

    # A. STRICT ROOM ASSIGNMENTS
    for c_idx, c in enumerate(data.courses):
        for r_idx, room in enumerate(data.rooms):
            is_lab_room = "lab" in room.lower()
            if c.is_lab != is_lab_room:
                for d in range(data.num_days):
                    for p in range(data.num_periods):
                        model.Add(schedule[(c_idx, r_idx, d, p)] == 0)

    # B. EARLY START (Only if the day is active)
    for section in data.sections:
        c_indices = [i for i, c in enumerate(data.courses) if c.section == section]
        for d in range(data.num_days):
            class_at_0 = sum(schedule[(i, r_idx, d, 0)] for i in c_indices for r_idx in range(len(data.rooms)))
            class_at_1 = sum(schedule[(i, r_idx, d, 1)] for i in c_indices for r_idx in range(len(data.rooms)))
            
            b0 = model.NewBoolVar(f'sec_{section}_d{d}_has_class_0')
            b1 = model.NewBoolVar(f'sec_{section}_d{d}_has_class_1')
            
            model.Add(class_at_0 > 0).OnlyEnforceIf(b0)
            model.Add(class_at_0 == 0).OnlyEnforceIf(b0.Not())
            model.Add(class_at_1 > 0).OnlyEnforceIf(b1)
            model.Add(class_at_1 == 0).OnlyEnforceIf(b1.Not())
            
            model.Add(b0 + b1 >= 1).OnlyEnforceIf(sec_active_day[(section, d)])

    # C. RECESS & ANTI-FATIGUE (Only if the day is active)
    for section in data.sections:
        c_indices = [i for i, c in enumerate(data.courses) if c.section == section]
        for d in range(data.num_days):
            model.Add(recess[(section, d, 2)] + recess[(section, d, 3)] == 1).OnlyEnforceIf(sec_active_day[(section, d)])
            model.Add(recess[(section, d, 2)] + recess[(section, d, 3)] == 0).OnlyEnforceIf(sec_active_day[(section, d)].Not())

            for p in range(data.num_periods):
                if p not in [2, 3]:
                    model.Add(recess[(section, d, p)] == 0)

            active_periods_vars = []
            for p in range(data.num_periods):
                period_active = model.NewBoolVar(f"sec_{section}_d{d}_p{p}_active_cutoff")
                total_classes_this_period = sum(
                    schedule[(i, r_idx, d, p)] for i in c_indices for r_idx in range(len(data.rooms))
                )
                
                model.Add(total_classes_this_period == 0).OnlyEnforceIf(recess[(section, d, p)])
                model.Add(total_classes_this_period > 0).OnlyEnforceIf(period_active)
                model.Add(total_classes_this_period == 0).OnlyEnforceIf(period_active.Not())
                active_periods_vars.append(period_active)

            for p in range(6, data.num_periods):
                model.Add(active_periods_vars[p] == 0).OnlyEnforceIf(recess[(section, d, 2)])
            for p in range(7, data.num_periods):
                model.Add(active_periods_vars[p] == 0).OnlyEnforceIf(recess[(section, d, 3)])

    # D. COURSE DELIVERY & CONTINUOUS LABS
    for c_idx, c in enumerate(data.courses):
        if not c.is_lab:
            model.Add(sum(
                schedule[(c_idx, r_idx, d, p)] 
                for r_idx in range(len(data.rooms)) for d in range(data.num_days) for p in range(data.num_periods)
            ) == c.hours)
        else:
            lab_start = {}
            for r_idx in range(len(data.rooms)):
                for d in range(data.num_days):
                    for p in range(data.num_periods - 1):
                        lab_start[(r_idx, d, p)] = model.NewBoolVar(f"ls_{c_idx}_r{r_idx}_d{d}_p{p}")

            model.AddExactlyOne(lab_start.values())

            for r_idx in range(len(data.rooms)):
                for d in range(data.num_days):
                    for p in range(data.num_periods):
                        starts_covering = []
                        if p < data.num_periods - 1: starts_covering.append(lab_start[(r_idx, d, p)])
                        if p > 0: starts_covering.append(lab_start[(r_idx, d, p-1)])
                        model.Add(schedule[(c_idx, r_idx, d, p)] == sum(starts_covering))

    # E. PARALLEL ELECTIVES
    for section in data.sections:
        c_indices = [i for i, c in enumerate(data.courses) if c.section == section]
        groups = {}
        for c_idx in c_indices:
            c = data.courses[c_idx]
            grp = c.elective_group if c.elective_group else f"core_{c_idx}"
            if grp not in groups: groups[grp] = []
            groups[grp].append(c_idx)

        for d in range(data.num_days):
            for p in range(data.num_periods):
                group_active_vars = []
                for grp_name, course_indices in groups.items():
                    grp_active = model.NewBoolVar(f"sec_{section}_grp_{grp_name}_d{d}_p{p}")
                    group_active_vars.append(grp_active)
                    for c_idx in course_indices:
                        c_active = sum(schedule[(c_idx, r_idx, d, p)] for r_idx in range(len(data.rooms)))
                        model.Add(c_active == grp_active)
                model.AddAtMostOne(group_active_vars)

    # F. MUTUAL EXCLUSIVITY (Rooms & Teachers)
    for d in range(data.num_days):
        for p in range(data.num_periods):
            for r_idx in range(len(data.rooms)):
                model.AddAtMostOne(schedule[(c_idx, r_idx, d, p)] for c_idx in range(len(data.courses)))
            for teacher in data.teachers:
                teacher_c_indices = [i for i, c in enumerate(data.courses) if teacher in c.teachers]
                model.AddAtMostOne(schedule[(c_idx, r_idx, d, p)] for c_idx in teacher_c_indices for r_idx in range(len(data.rooms)))

    # G. MAX 1 LAB GROUP PER DAY
    for section in data.sections:
        c_indices = [i for i, c in enumerate(data.courses) if c.section == section]
        lab_groups = {}
        for c_idx in c_indices:
            c = data.courses[c_idx]
            if c.is_lab:
                grp_name = c.elective_group if c.elective_group else f"core_lab_{c_idx}"
                if grp_name not in lab_groups: lab_groups[grp_name] = []
                lab_groups[grp_name].append(c_idx)

        for d in range(data.num_days):
            daily_active_lab_groups = []
            for grp_name, c_grp_indices in lab_groups.items():
                grp_active_today = model.NewBoolVar(f"lab_grp_{grp_name}_active_d{d}")
                total_classes_today = sum(
                    schedule[(i, r_idx, d, p)] 
                    for i in c_grp_indices for r_idx in range(len(data.rooms)) for p in range(data.num_periods)
                )
                model.Add(total_classes_today > 0).OnlyEnforceIf(grp_active_today)
                model.Add(total_classes_today == 0).OnlyEnforceIf(grp_active_today.Not())
                daily_active_lab_groups.append(grp_active_today)
            model.Add(sum(daily_active_lab_groups) <= 1)

    # H. LOAD BALANCING (Relaxed to handle Day Offs)
    for section in data.sections:
        c_indices = [i for i, c in enumerate(data.courses) if c.section == section]
        unique_groups_hours = {}
        for c_idx in c_indices:
            c = data.courses[c_idx]
            grp = c.elective_group if c.elective_group else f"core_{c_idx}"
            if grp not in unique_groups_hours: unique_groups_hours[grp] = c.hours
        
        total_student_hours = sum(unique_groups_hours.values())
        avg_hours = total_student_hours / data.num_days
        max_classes = min(data.num_periods - 1, int(math.ceil(avg_hours)) + 2)

        for d in range(data.num_days):
            active_periods_today = []
            for p in range(data.num_periods):
                period_active = model.NewBoolVar(f"sec_{section}_d{d}_p{p}_active_load")
                total_classes_this_period = sum(
                    schedule[(i, r_idx, d, p)] for i in c_indices for r_idx in range(len(data.rooms))
                )
                model.Add(total_classes_this_period > 0).OnlyEnforceIf(period_active)
                model.Add(total_classes_this_period == 0).OnlyEnforceIf(period_active.Not())
                active_periods_today.append(period_active)
            model.Add(sum(active_periods_today) <= max_classes)

    # I. CONTIGUOUS SCHEDULING
    for section in data.sections:
        c_indices = [i for i, c in enumerate(data.courses) if c.section == section]
        for d in range(data.num_days):
            is_occupied = []
            for p in range(data.num_periods):
                occupied_var = model.NewBoolVar(f"sec_{section}_d{d}_p{p}_occupied")
                total_events = sum(
                    schedule[(i, r_idx, d, p)] 
                    for i in c_indices for r_idx in range(len(data.rooms))
                ) + recess[(section, d, p)]
                
                model.Add(total_events > 0).OnlyEnforceIf(occupied_var)
                model.Add(total_events == 0).OnlyEnforceIf(occupied_var.Not())
                is_occupied.append(occupied_var)
                
            transitions = []
            for p in range(data.num_periods):
                t = model.NewBoolVar(f"sec_{section}_d{d}_p{p}_transition")
                if p == 0:
                    model.Add(t == is_occupied[0])
                else:
                    model.Add(t >= is_occupied[p] - is_occupied[p-1])
                transitions.append(t)
            
            model.Add(sum(transitions) <= 1).OnlyEnforceIf(sec_active_day[(section, d)])

    # J. FACULTY WORKLOAD
    max_teacher_hours_per_day = 4
    for teacher in data.teachers:
        teacher_c_indices = [i for i, c in enumerate(data.courses) if teacher in c.teachers]
        if not teacher_c_indices: continue
            
        for d in range(data.num_days):
            daily_teacher_classes = []
            for p in range(data.num_periods):
                for r_idx in range(len(data.rooms)):
                    for c_idx in teacher_c_indices:
                        daily_teacher_classes.append(schedule[(c_idx, r_idx, d, p)])
            model.Add(sum(daily_teacher_classes) <= max_teacher_hours_per_day)

    # K. SUBJECT FATIGUE
    for c_idx, c in enumerate(data.courses):
        if not c.is_lab:
            for d in range(data.num_days):
                model.Add(sum(
                    schedule[(c_idx, r_idx, d, p)] 
                    for r_idx in range(len(data.rooms)) for p in range(data.num_periods)
                ) <= 2)

    # --- SOLVE ---
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 45.0 
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        result = {section: [] for section in data.sections}
        for section in data.sections:
            for d in range(data.num_days):
                for p in range(data.num_periods):
                    if solver.Value(recess[(section, d, p)]) == 1:
                        result[section].append(ScheduledClass(
                            day=d, period=p + 1, room="Campus", course_name="RECESS", teachers=[], is_recess=True
                        ))
                    else:
                        for r_idx, room in enumerate(data.rooms):
                            for c_idx, c in enumerate(data.courses):
                                if c.section == section and solver.Value(schedule[(c_idx, r_idx, d, p)]) == 1:
                                    result[section].append(ScheduledClass(
                                        day=d, period=p + 1, room=room, course_name=c.name, teachers=c.teachers
                                    ))
        return {"status": "success", "schedule": result, "message": "Timetable generated successfully"}
    else:
        return {"status": "failed", "schedule": None, "message": "Could not find a conflict-free timetable."}

# ==========================================
# 3. FASTAPI ENDPOINTS
# ==========================================

@app.post("/api/v1/generate", response_model=TimetableResponse)
async def create_timetable(request: TimetableRequest):
    try:
        result = generate_schedule(request)
        if result["status"] == "failed":
            raise HTTPException(status_code=400, detail=result["message"])
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/reschedule-dynamic", response_model=TimetableResponse)
async def dynamic_reschedule(request: RescheduleRequest):
    schedule = request.current_schedule
    section = request.course.section
    teachers = request.course.teachers
    is_course_lab = request.course.is_lab 
    
    for p in range(1, request.num_periods + 1):
        if any(c.day == request.target_day and c.period == p for c in schedule.get(section, [])): continue
            
        teacher_busy = False
        for sec, classes in schedule.items():
            for c in classes:
                if c.day == request.target_day and c.period == p and any(t in c.teachers for t in teachers):
                    teacher_busy = True
                    break
            if teacher_busy: break
        if teacher_busy: continue
            
        occupied = [c.room for sec, classes in schedule.items() for c in classes if c.day == request.target_day and c.period == p and not c.is_recess]
        
        available = []
        for r in request.all_rooms:
            is_lab_room = "lab" in r.lower()
            if r not in occupied and is_course_lab == is_lab_room:
                available.append(r)
        
        if available:
            new_class = ScheduledClass(
                day=request.target_day, period=p, room=available[0],
                course_name=request.course.name + " (Rescheduled)", teachers=teachers
            )
            if section not in schedule: schedule[section] = []
            schedule[section].append(new_class)
            schedule[section] = sorted(schedule[section], key=lambda x: (x.day, x.period))
            return {"status": "success", "schedule": schedule, "message": f"Rescheduled to Period {p} in {available[0]}."}
            
    raise HTTPException(status_code=400, detail="No available slots found in matching room types on this day.")

@app.post("/api/v1/schedule-extra", response_model=TimetableResponse)
async def schedule_extra_class(request: ExtraClassRequest):
    """Greedy search across the week to find an empty slot for an extra class."""
    schedule = request.current_schedule
    section = request.section
    
    # 1. Fetch course details to know the teachers and lab status
    target_course = next((c for c in request.all_courses if c.id == request.course_id and c.section == section), None)
    
    if not target_course:
        raise HTTPException(status_code=404, detail="Course ID not found for the provided section.")
        
    teachers = target_course.teachers
    is_course_lab = target_course.is_lab
    
    # Determine the days to search. If user provided a target day, search only that. Otherwise, search all.
    days_to_search = [request.target_day] if request.target_day is not None else range(request.num_days)
    
    for d in days_to_search:
        # Periods are 1-indexed in our JSON (1 to 8)
        for p in range(1, request.num_periods + 1):
            
            # Check 1: Is the section busy on this day & period?
            section_busy = any(c.day == d and c.period == p for c in schedule.get(section, []))
            if section_busy:
                continue
                
            # Check 2: Are the required teachers busy in ANY other section?
            teacher_busy = False
            for sec, classes in schedule.items():
                for c in classes:
                    if c.day == d and c.period == p and any(t in c.teachers for t in teachers):
                        teacher_busy = True
                        break
                if teacher_busy: break
                
            if teacher_busy:
                continue
                
            # Check 3: Room Availability & Strict Lab Matching
            occupied_rooms = [
                c.room for sec, classes in schedule.items() for c in classes 
                if c.day == d and c.period == p and not c.is_recess
            ]
            
            available_rooms = []
            for r in request.all_rooms:
                is_lab_room = "lab" in r.lower()
                # Must be empty AND must match the course type (lab -> lab, theory -> theory)
                if r not in occupied_rooms and is_course_lab == is_lab_room:
                    available_rooms.append(r)
            
            if available_rooms:
                # Success! We found a slot.
                new_class = ScheduledClass(
                    day=d, period=p, room=available_rooms[0],
                    course_name=target_course.name + " (Extra)", teachers=teachers
                )
                
                if section not in schedule: schedule[section] = []
                schedule[section].append(new_class)
                
                # Keep the schedule chronologically sorted
                schedule[section] = sorted(schedule[section], key=lambda x: (x.day, x.period))
                
                day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
                return {
                    "status": "success", 
                    "schedule": schedule, 
                    "message": f"Extra class scheduled successfully on {day_names[d]}, Period {p} in {available_rooms[0]}."
                }
                
    raise HTTPException(status_code=400, detail="Could not find a conflict-free slot for this extra class. All teachers or suitable rooms are busy.")


# ==========================================
# 4. EXPORT TO EXCEL MODELS & ENDPOINT (Beautiful Format)
# ==========================================

@app.post("/api/v1/export/excel")
async def export_schedule_to_excel(request: ExportRequest):
    try:
        wb = Workbook()
        
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        day_names = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]
        times = ["9:00 - 10:00", "10:00 - 11:00", "11:00 - 12:00", "12:00 - 1:00", "1:00 - 2:00", "2:00 - 3:00", "3:00 - 4:00", "4:00 - 5:00"]
        
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

        for section, classes in request.schedule.items():
            ws = wb.create_sheet(title=section)
            
            ws.column_dimensions['A'].width = 15
            for col in range(2, request.num_periods + 2):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24

            ws.merge_cells('A1:C1')
            ws['A1'] = f"Shift/Semester: {section}"
            ws['A1'].font = bold_font
            
            ws.merge_cells('D1:F1')
            ws['D1'] = "Session: 2025-26"
            ws['D1'].font = bold_font
            ws['D1'].alignment = Alignment(horizontal="center")
            
            ws.merge_cells('G1:I1')
            from datetime import date
            ws['G1'] = f"Date: {date.today().strftime('%d/%m/%Y')}"
            ws['G1'].font = bold_font
            ws['G1'].alignment = Alignment(horizontal="right")

            ws['A2'] = "Lecture"
            ws['A3'] = "TIME\nDAY"
            ws.row_dimensions[3].height = 30
            
            for i in range(request.num_periods):
                ws.cell(row=2, column=i+2, value=str(i+1))
                ws.cell(row=3, column=i+2, value=times[i] if i < len(times) else "---")

            for i in range(request.num_days):
                ws.cell(row=i+4, column=1, value=day_names[i])
                ws.row_dimensions[i+4].height = 80 

            unique_courses = {}
            
            # --- AGGREGATE CLASSES BY SLOT ---
            grid = {}
            for c in classes:
                if c.day >= request.num_days or c.period > request.num_periods:
                    continue
                key = (c.day, c.period)
                if key not in grid:
                    grid[key] = []
                grid[key].append(c)

            # --- POPULATE THE GRID ---
            for (day, period), slot_classes in grid.items():
                row = day + 4
                col = period + 1
                cell = ws.cell(row=row, column=col)
                
                if len(slot_classes) == 1 and slot_classes[0].is_recess:
                    cell.value = "RECESS"
                    cell.font = bold_font
                else:
                    cell_texts = []
                    for c in slot_classes:
                        teachers_str = ", ".join(c.teachers)
                        cell_texts.append(f"{c.course_name}\n({teachers_str})\n[{c.room}]")
                        unique_courses[c.course_name] = teachers_str
                    
                    cell.value = "\n--- OR ---\n".join(cell_texts)

            for r in range(2, request.num_days + 4):
                for col_idx in range(1, request.num_periods + 2):
                    cell = ws.cell(row=r, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = center_aligned
                    if r <= 3 or col_idx == 1:
                        cell.fill = header_fill
                        cell.font = bold_font

            legend_start_row = request.num_days + 6
            
            ws.cell(row=legend_start_row, column=2, value="Time Table In-Charge").font = bold_font
            ws.cell(row=legend_start_row, column=request.num_periods, value="HOD").font = bold_font
            
            legend_table_row = legend_start_row + 2
            ws.cell(row=legend_table_row, column=2, value="Course Details").font = bold_font
            ws.cell(row=legend_table_row, column=3, value="Course Coordinator").font = bold_font
            
            ws.cell(row=legend_table_row, column=2).border = thin_border
            ws.cell(row=legend_table_row, column=3).border = thin_border
            ws.cell(row=legend_table_row, column=2).fill = header_fill
            ws.cell(row=legend_table_row, column=3).fill = header_fill
            
            current_row = legend_table_row + 1
            for course, coordinator in unique_courses.items():
                if course == "RECESS": continue
                ws.cell(row=current_row, column=2, value=course).border = thin_border
                ws.cell(row=current_row, column=3, value=coordinator).border = thin_border
                current_row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        headers = {'Content-Disposition': 'attachment; filename="AcadFlow_Timetable_Formatted.xlsx"'}
        return StreamingResponse(
            output, 
            headers=headers, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {str(e)}")